import hashlib
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MASTER_SHEET_NAME = "Giá nhập"
HISTORY_SHEET_NAME = "Lịch sử nhập hàng"

MASTER_HEADERS = (
    "Sản phẩm",
    "Giá nhập",
)

HISTORY_HEADERS = (
    "Tên sản phẩm",
    "Số lượng được duyệt",
    "Ngày nhập hàng",
    "Ngày về kho",
    "Chi phí nhập thực tế/ 1 sp",
)


@dataclass(frozen=True)
class ProductCostWorkbook:
    source_file_name: str
    source_file_sha256: bytes
    source_file_size_bytes: int
    source_file_modified_at_utc: datetime
    master_sheet_name: str
    history_sheet_name: str
    master_rows: list[dict[str, Any]]
    history_rows: list[dict[str, Any]]


def _file_sha256(file_path: Path) -> bytes:
    digest = hashlib.sha256()

    with file_path.open("rb") as file:
        for chunk in iter(
            lambda: file.read(1024 * 1024),
            b"",
        ):
            digest.update(chunk)

    return digest.digest()


def _clean_text(value: Any) -> str | None:
    if value in (None, ""):
        return None

    text = str(value).strip()
    return text or None


def _validate_headers(
    worksheet: Any,
    expected_headers: tuple[str, ...],
) -> None:
    actual_headers = tuple(
        _clean_text(
            worksheet.cell(
                row=1,
                column=column_number,
            ).value
        )
        for column_number in range(
            1,
            len(expected_headers) + 1,
        )
    )

    if actual_headers != expected_headers:
        raise ValueError(
            f"Unexpected headers in sheet "
            f"{worksheet.title!r}. "
            f"Expected {expected_headers!r}, "
            f"received {actual_headers!r}."
        )


def extract_product_cost_workbook(
    file_path: str | Path,
) -> ProductCostWorkbook:
    resolved_path = Path(file_path).resolve(strict=True)

    if resolved_path.suffix.lower() != ".xlsx":
        raise ValueError(
            "Product cost source must be an .xlsx file."
        )

    file_stat = resolved_path.stat()

    workbook = load_workbook(
        resolved_path,
        read_only=True,
        data_only=True,
    )

    try:
        required_sheets = {
            MASTER_SHEET_NAME,
            HISTORY_SHEET_NAME,
        }
        missing_sheets = required_sheets.difference(
            workbook.sheetnames
        )

        if missing_sheets:
            raise ValueError(
                "Workbook is missing required sheets: "
                + ", ".join(sorted(missing_sheets))
            )

        master_sheet = workbook[MASTER_SHEET_NAME]
        history_sheet = workbook[HISTORY_SHEET_NAME]

        _validate_headers(
            master_sheet,
            MASTER_HEADERS,
        )
        _validate_headers(
            history_sheet,
            HISTORY_HEADERS,
        )

        master_rows: list[dict[str, Any]] = []

        for row_number, row in enumerate(
            master_sheet.iter_rows(
                min_row=2,
                max_col=2,
                values_only=True,
            ),
            start=2,
        ):
            product_name = _clean_text(row[0])

            if product_name is None:
                continue

            master_rows.append(
                {
                    "source_row_number": row_number,
                    "product_name": product_name,
                    "unit_cost": row[1],
                }
            )

        history_rows: list[dict[str, Any]] = []

        for row_number, row in enumerate(
            history_sheet.iter_rows(
                min_row=2,
                max_col=5,
                values_only=True,
            ),
            start=2,
        ):
            product_name = _clean_text(row[0])

            if product_name is None:
                continue

            history_rows.append(
                {
                    "source_row_number": row_number,
                    "product_name": product_name,
                    "approved_quantity": row[1],
                    "import_date": row[2],
                    "warehouse_arrival_date": row[3],
                    "actual_unit_cost": row[4],
                }
            )
    finally:
        workbook.close()

    return ProductCostWorkbook(
        source_file_name=resolved_path.name,
        source_file_sha256=_file_sha256(
            resolved_path
        ),
        source_file_size_bytes=file_stat.st_size,
        source_file_modified_at_utc=datetime.fromtimestamp(
            file_stat.st_mtime,
            tz=UTC,
        ).replace(tzinfo=None),
        master_sheet_name=MASTER_SHEET_NAME,
        history_sheet_name=HISTORY_SHEET_NAME,
        master_rows=master_rows,
        history_rows=history_rows,
    )